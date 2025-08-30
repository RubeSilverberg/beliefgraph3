#!/usr/bin/env python3
"""
Excel → Minimal JSON (Cell-level Lineage Subgraph)

Creates a compact provenance graph:
  - Pick top-N formula cells from projection-like sheets (by inbound cross-sheet refs).
  - For each target formula cell, add edges from referenced source cells (cross-sheet only).
  - Include helpful descriptions: formula text, source column header, row label.

Output: tests/minimal-json/excel-lineage-deps.json

Usage:
  python3 tests/dev-tools/excel_lineage_to_minimal.py "Field Goal 1 Projections (1).xlsx" [N]
Where N (optional) is the max number of target cells to include (default 8).
"""

import json
import re
import sys
import unicodedata
from collections import defaultdict, Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    print("Missing dependency 'openpyxl'. Install with: pip install openpyxl", file=sys.stderr)
    raise


CELL_REF_RE = re.compile(r"(?:(?:'([^']+)')|([A-Za-z0-9_]+))!\$?([A-Za-z]{1,3})\$?(\d+)")


def ascii_id(s: str) -> str:
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^A-Za-z0-9_.:-]+", "_", s).strip("_")
    return s[:160] if len(s) > 160 else s


def col_to_index(col_letters: str) -> int:
    idx = 0
    for ch in col_letters.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


def index_to_col(idx: int) -> str:
    letters = ''
    n = idx
    while n:
        n, rem = divmod(n-1, 26)
        letters = chr(65 + rem) + letters
    return letters or 'A'


def find_column_header(ws, col_idx: int, max_scan_rows: int = 10) -> str:
    for r in range(1, max_scan_rows + 1):
        val = ws.cell(row=r, column=col_idx).value
        if isinstance(val, str) and val.strip():
            return val.strip()
    return f"Col {index_to_col(col_idx)}"


def find_row_label(ws, row_idx: int, max_scan_cols: int = 3) -> str:
    # Heuristic: look in first few columns for a string label
    for c in range(1, max_scan_cols + 1):
        val = ws.cell(row=row_idx, column=c).value
        if isinstance(val, str) and val.strip():
            return val.strip()
    return f"Row {row_idx}"


def detect_projection_sheets(wb):
    # Strategy: sheets with many formulas and many cross-sheet references
    formula_counts = Counter()
    inbound_xsheet = Counter()
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    formula_counts[sname] += 1
                    for m in CELL_REF_RE.finditer(val):
                        ref_sheet = (m.group(1) or m.group(2) or '').strip()
                        if ref_sheet and ref_sheet != sname:
                            inbound_xsheet[sname] += 1
    # Rank by inbound cross-sheet refs
    ranked = [s for s, _ in inbound_xsheet.most_common()]
    # Ensure obvious names (e.g., contain 'Projection') bubble up even if ties
    ranked = sorted(ranked, key=lambda n: (('projection' not in n.lower()), -inbound_xsheet[n], -formula_counts[n]))
    # Keep top 2–3
    return ranked[:3]


def build_lineage_graph(xlsx_path: Path, max_targets: int = 8, max_refs_per_target: int = 6):
    wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=True)
    target_sheets = detect_projection_sheets(wb)
    nodes = []
    edges = []
    node_ids = set()

    # Create nodes for target sheets
    for tname in target_sheets:
        if tname not in node_ids:
            nodes.append({
                'id': tname,
                'label': tname[:22] + ('…' if len(tname) > 22 else ''),
                'type': 'assertion',
                'description': f"Projection sheet '{tname}'."
            })
            node_ids.add(tname)

    # Collect candidate target cells
    candidates = []  # (tname, cell_addr, refs, formula)
    for tname in target_sheets:
        ws_t = wb[tname]
        for row in ws_t.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    refs = [m.groups() for m in CELL_REF_RE.finditer(val)]
                    xs_refs = [(rs or rp, col, int(rn)) for (rs, rp, col, rn) in refs if (rs or rp) and (rs or rp) != tname]
                    if xs_refs:
                        addr = f"{tname}!{cell.column_letter}{cell.row}"
                        candidates.append((tname, addr, xs_refs, val))

    # Rank targets by number of cross-sheet refs and pick top-N
    candidates.sort(key=lambda x: len(x[2]), reverse=True)
    picks = candidates[:max_targets]

    # Build nodes and edges for each picked target cell
    for (tname, addr, xs_refs, formula) in picks:
        # Target cell node
        tnode_id = ascii_id(addr)
        if tnode_id not in node_ids:
            nodes.append({
                'id': tnode_id,
                'label': addr.split('!')[-1],  # show just A1 for compactness
                'type': 'assertion',
                'description': f"Target cell {addr}\nFormula: {formula}"
            })
            node_ids.add(tnode_id)
        # Link target cell to its sheet (so sheets remain the hub)
        edges.append({
            'source': tnode_id,
            'target': tname,
            'type': 'supports',
            'weight': 0.6,
            'contributingFactors': ["Selected output cell"]
        })

        # Compute per-source weights normalized per target cell
        per = max(1, len(xs_refs))
        w_each = round(1.0 / per, 4)

        # Add source cell nodes and edges
        # Limit to first M refs to keep the graph small
        for (sname, col_letters, row_num) in xs_refs[:max_refs_per_target]:
            ws_s = wb[sname]
            col_idx = col_to_index(col_letters)
            header = find_column_header(ws_s, col_idx)
            row_label = find_row_label(ws_s, row_num)
            src_addr = f"{sname}!{col_letters}{row_num}"
            snode_id = ascii_id(src_addr)
            if snode_id not in node_ids:
                nodes.append({
                    'id': snode_id,
                    'label': f"{col_letters}{row_num}",
                    'type': 'fact',
                    'description': f"Source {src_addr}\nHeader: {header}\nRow: {row_label}"
                })
                node_ids.add(snode_id)
            edges.append({
                'source': snode_id,
                'target': tnode_id,
                'type': 'supports',
                'weight': w_each,
                'contributingFactors': [f"Header: {header}", f"Row: {row_label}"]
            })

    return {
        'version': '2',
        'nodes': nodes,
        'edges': edges,
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: excel_lineage_to_minimal.py <workbook.xlsx> [max_targets] [output.json]", file=sys.stderr)
        sys.exit(2)
    xlsx = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx.exists():
        print(f"File not found: {xlsx}", file=sys.stderr)
        sys.exit(1)
    max_targets = int(sys.argv[2]) if len(sys.argv) > 2 and sys.argv[2].isdigit() else 8
    out = Path(sys.argv[3]).expanduser().resolve() if len(sys.argv) > 3 else Path(__file__).parent.parent / 'minimal-json' / 'excel-lineage-deps.json'
    out.parent.mkdir(parents=True, exist_ok=True)

    minimal = build_lineage_graph(xlsx, max_targets=max_targets)
    with out.open('w', encoding='utf-8') as f:
        json.dump(minimal, f, ensure_ascii=False, indent=2)

    print(f"Wrote minimal JSON: {out}")
    print(f"Nodes: {len(minimal['nodes'])}, Edges: {len(minimal['edges'])}")


if __name__ == '__main__':
    main()
