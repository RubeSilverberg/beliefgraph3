#!/usr/bin/env python3
"""
Excel → Minimal JSON (Sheet Dependency Graph)

Purpose
-------
Low-effort, native value-add for the Belief Graph tool: convert a workbook
into a causal-style sheet dependency graph by scanning formulas. Each sheet
becomes a node; if a formula in SheetB references cells in SheetA, we add
an edge A -> B. Edge weight is proportional to reference counts (0..1).

Output
------
Writes a minimal JSON file you can load in the existing converter/test page:
  tests/minimal-json/excel-sheet-deps.json

How to use
----------
- Ensure Python 3 is available. Install dependency:
    pip install openpyxl
- Run:
    python3 tests/dev-tools/excel_to_minimal.py "Field Goal 1 Projections (1).xlsx"
- Then open in browser:
    tests/minimal-json/test-minimal-converter.html
  and use the "Load Minimal JSON File" section to load excel-sheet-deps.json,
  then "Convert to Full" and optionally open the main app.

Notes
-----
- We do not evaluate formulas; we just parse text to find Sheet!Cell refs.
- This produces a compact, explainable graph unique to this workbook.
"""

import json
import re
import sys
from collections import defaultdict, Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as e:  # pragma: no cover
    print("Missing dependency 'openpyxl'. Install with: pip install openpyxl", file=sys.stderr)
    raise


REF_RE = re.compile(r"(?:(?:'([^']+)')|([A-Za-z0-9_]+))!\$?[A-Za-z]{1,3}\$?\d+", re.UNICODE)


def find_sheet_refs(formula: str):
    """Return a list of sheet names referenced in a formula string.
    Handles both 'Sheet Name'!A1 and SheetName!A1 patterns.
    """
    if not formula or not isinstance(formula, str):
        return []
    refs = []
    for m in REF_RE.finditer(formula):
        sheet_quoted, sheet_plain = m.groups()
        name = (sheet_quoted or sheet_plain or '').strip()
        if name:
            refs.append(name)
    return refs


def build_sheet_dependency_graph(xlsx_path: Path):
    wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=True)
    sheet_names = wb.sheetnames
    # Map: target_sheet -> Counter(source_sheet -> count)
    inbound = {name: Counter() for name in sheet_names}
    formula_counts = Counter()

    for sname in sheet_names:
        ws = wb[sname]
        # Iterate cells; for read_only, iterate_rows yields Cell with .value and .data_type
        for row in ws.iter_rows():
            for cell in row:
                # In openpyxl, formulas have data_type 'f' or value starting with '='
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    formula_counts[sname] += 1
                    refs = find_sheet_refs(val)
                    for ref_sheet in refs:
                        if ref_sheet in inbound:  # only count known sheets
                            inbound[sname][ref_sheet] += 1

    # Build minimal JSON
    nodes = []
    edges = []

    # Create nodes with provenance-rich description
    for sname in sheet_names:
        total_formulas = formula_counts.get(sname, 0)
        inbound_refs = sum(inbound[sname].values())
        top_sources = inbound[sname].most_common(3)
        top_str = ', '.join(f"{src} ({cnt})" for src, cnt in top_sources) if top_sources else 'None'
        desc = (
            f"Sheet '{sname}'. Contains {total_formulas} formulas; "
            f"inbound cross-sheet refs: {inbound_refs}. Top sources: {top_str}."
        )
        nodes.append({
            "id": sname,
            "label": sname[:22] + ('…' if len(sname) > 22 else ''),
            "type": "assertion",  # sheets act like computed assertions if they have inputs
            "description": desc
        })

    # Normalize edge weights per target (sum inbounds → 1.0)
    for target, sources in inbound.items():
        total = sum(sources.values())
        if total <= 0:
            continue
        for source, cnt in sources.items():
            if source == target:
                # Skip self loops; they aren't helpful at sheet level
                continue
            w = cnt / total
            edges.append({
                "source": source,
                "target": target,
                "type": "supports",
                "weight": round(min(max(w, 0.0), 1.0), 4),
                "contributingFactors": [
                    f"{cnt} cross-sheet reference(s)"
                ]
            })

    minimal = {
        "version": "2",
        "nodes": nodes,
        "edges": edges,
    }
    return minimal


def main():
    if len(sys.argv) < 2:
        print("Usage: excel_to_minimal.py <workbook.xlsx> [output.json]", file=sys.stderr)
        sys.exit(2)
    xlsx = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx.exists():
        print(f"File not found: {xlsx}", file=sys.stderr)
        sys.exit(1)
    out = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) > 2 else Path(__file__).parent.parent / 'minimal-json' / 'excel-sheet-deps.json'
    out.parent.mkdir(parents=True, exist_ok=True)

    minimal = build_sheet_dependency_graph(xlsx)
    with out.open('w', encoding='utf-8') as f:
        json.dump(minimal, f, ensure_ascii=False, indent=2)

    # Console summary
    print(f"Wrote minimal JSON: {out}")
    print(f"Nodes: {len(minimal['nodes'])}, Edges: {len(minimal['edges'])}")


if __name__ == '__main__':
    main()
