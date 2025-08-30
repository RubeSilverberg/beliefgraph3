#!/usr/bin/env python3
"""
Excel → Minimal JSON (Column/Topic Dependency Graph)

Builds a more informative graph than sheet→sheet by grouping cross-sheet
references by their source column "topic" (header near the referenced cell).

For each formula in a target sheet, we parse referenced cells. For any
cross-sheet reference (SheetA!ColRow), we look up the header text for that
column on SheetA (by scanning the top 10 rows in the same column) and treat
that as a topic node: "SheetA::Header". We then add an edge Topic → TargetSheet.

Output: tests/minimal-json/excel-topics-deps.json

Usage:
  python3 tests/dev-tools/excel_topics_to_minimal.py "Field Goal 1 Projections (1).xlsx"

Notes:
  - No formula evaluation, just lexical parsing; fast + safe.
  - Works even when headers are multiline or spaced; falls back to "Col X".
  - Keeps IDs ASCII-safe; labels are human-friendly.
"""

import json
import re
import sys
import unicodedata
from collections import defaultdict, Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    print("Missing dependency 'openpyxl'. Install with: pip install openpyxl", file=sys.stderr)
    raise


CELL_REF_RE = re.compile(r"(?:(?:'([^']+)')|([A-Za-z0-9_]+))!\$?([A-Za-z]{1,3})\$?(\d+)")


def ascii_id(s: str) -> str:
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^A-Za-z0-9_.:-]+", "_", s).strip("_")
    return s[:120] if len(s) > 120 else s


def col_to_index(col_letters: str) -> int:
    idx = 0
    for ch in col_letters.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


def find_column_header(ws, col_idx: int, max_scan_rows: int = 10) -> str:
    """Find a likely header for a column by scanning top rows in that column.
    Returns first non-empty text cell (stripped). Fallback: "Col <letter>".
    """
    header = None
    for r in range(1, max_scan_rows + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = cell.value
        if isinstance(val, str):
            t = val.strip()
            if t:
                header = t
                break
        elif isinstance(val, (int, float)):
            # Numeric isn't a great header; keep looking
            continue
    if not header:
        # Fallback to A,B,C name
        # Convert index to letters
        n = col_idx
        letters = ''
        while n:
            n, rem = divmod(n-1, 26)
            letters = chr(65 + rem) + letters
        header = f"Col {letters}"
    return header


def build_topic_graph(xlsx_path: Path):
    wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=True)
    sheet_names = wb.sheetnames

    # Count: (topic_id, target_sheet) -> references
    topic_to_target = Counter()
    # Map topic_id -> { sheet, header, sample_cells:set }
    topic_meta = {}
    # Track simple sheet nodes for targets
    target_formula_counts = Counter()

    for tname in sheet_names:
        ws_t = wb[tname]
        for row in ws_t.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    target_formula_counts[tname] += 1
                    for m in CELL_REF_RE.finditer(val):
                        sname = (m.group(1) or m.group(2) or '').strip()
                        col_letters = m.group(3)
                        try:
                            row_num = int(m.group(4))
                        except Exception:
                            row_num = None
                        if not sname or sname == tname:
                            continue  # same-sheet or invalid
                        if sname not in wb.sheetnames:
                            continue
                        ws_s = wb[sname]
                        cidx = col_to_index(col_letters)
                        header = find_column_header(ws_s, cidx)
                        topic_id = ascii_id(f"{sname}::{header}")
                        if topic_id not in topic_meta:
                            topic_meta[topic_id] = {
                                'sheet': sname,
                                'header': header,
                                'samples': set()
                            }
                        # Sample up to a few examples for description
                        if row_num is not None and len(topic_meta[topic_id]['samples']) < 4:
                            topic_meta[topic_id]['samples'].add(f"{sname}!{col_letters}{row_num}")
                        topic_to_target[(topic_id, tname)] += 1

    # Build nodes
    nodes = []
    # Topic nodes
    for tid, meta in topic_meta.items():
        label = meta['header'][:42] + ('…' if len(meta['header']) > 42 else '')
        samples = sorted(meta['samples'])
        desc = (
            f"Topic from sheet '{meta['sheet']}': {meta['header']}\n" +
            (f"Examples: {', '.join(samples)}" if samples else "")
        ).strip()
        nodes.append({
            'id': tid,
            'label': label,
            'type': 'fact',  # act like inputs to targets
            'description': desc
        })
    # Target sheet nodes
    for tname, fcnt in target_formula_counts.items():
        nodes.append({
            'id': tname,
            'label': tname[:22] + ('…' if len(tname) > 22 else ''),
            'type': 'assertion',
            'description': f"Sheet '{tname}' with {fcnt} formulas referencing external topics"
        })

    # Build edges with per-target normalization
    # For each target, sum counts and compute weights
    counts_by_target = defaultdict(int)
    for (tid, tname), cnt in topic_to_target.items():
        counts_by_target[tname] += cnt
    edges = []
    for (tid, tname), cnt in topic_to_target.items():
        total = counts_by_target[tname] or 1
        w = round(min(max(cnt / total, 0.0), 1.0), 4)
        meta = topic_meta.get(tid, {})
        src_sheet = meta.get('sheet', '')
        edges.append({
            'source': tid,
            'target': tname,
            'type': 'supports',
            'weight': w,
            'contributingFactors': [
                f"{cnt} ref(s) from {src_sheet}::{meta.get('header','')}"
            ]
        })

    minimal = {
        'version': '2',
        'nodes': nodes,
        'edges': edges,
    }
    return minimal


def main():
    if len(sys.argv) < 2:
        print("Usage: excel_topics_to_minimal.py <workbook.xlsx> [output.json]", file=sys.stderr)
        sys.exit(2)
    xlsx = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx.exists():
        print(f"File not found: {xlsx}", file=sys.stderr)
        sys.exit(1)
    out = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) > 2 else Path(__file__).parent.parent / 'minimal-json' / 'excel-topics-deps.json'
    out.parent.mkdir(parents=True, exist_ok=True)

    minimal = build_topic_graph(xlsx)
    with out.open('w', encoding='utf-8') as f:
        json.dump(minimal, f, ensure_ascii=False, indent=2)

    print(f"Wrote minimal JSON: {out}")
    print(f"Nodes: {len(minimal['nodes'])}, Edges: {len(minimal['edges'])}")


if __name__ == '__main__':
    main()
